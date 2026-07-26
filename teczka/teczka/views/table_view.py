"""Table/data comparison view for CSV and Excel files."""

from __future__ import annotations

import csv
import os
from typing import Optional

from PySide6.QtCore import Qt
from PySide6.QtGui import QColor, QStandardItem, QStandardItemModel
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QSplitter,
    QTabWidget,
    QTableView,
    QVBoxLayout,
    QWidget,
)

from ..workers.function_worker import FunctionWorker

# ---------------------------------------------------------------------------
# Cell status constants and background colours
# ---------------------------------------------------------------------------

STATUS_IDENTICAL = "identical"
STATUS_DIFFERENT = "different"
STATUS_LEFT_ONLY = "left_only"
STATUS_RIGHT_ONLY = "right_only"
STATUS_MISSING = "missing"

_STATUS_COLORS: dict[str, QColor] = {
    STATUS_IDENTICAL: QColor("#c8e6c9"),   # green
    STATUS_DIFFERENT: QColor("#ffcdd2"),   # red
    STATUS_LEFT_ONLY: QColor("#fff9c4"),   # yellow
    STATUS_RIGHT_ONLY: QColor("#fff9c4"),  # yellow
    STATUS_MISSING: QColor("#f5f5f5"),     # gray
}


def _read_csv_rows(path: str) -> list[list[str]]:
    """Read a CSV file and return a list of rows (each a list of strings).

    Pure I/O + parsing (no Qt access) so it can run on a worker thread.
    """
    if not path or not os.path.isfile(path):
        return []
    try:
        with open(path, "r", newline="", errors="replace") as fh:
            # Sniff dialect for up to 8 KB
            sample = fh.read(8192)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel  # type: ignore[assignment]
            reader = csv.reader(fh, dialect)
            return [row for row in reader]
    except (OSError, csv.Error):
        return []


def _read_excel_sheets(path: str) -> dict[str, list[list[str]]]:
    """Read an Excel file and return {sheet_name: rows}.

    Pure I/O + parsing (no Qt access) so it can run on a worker thread.
    Requires openpyxl; raises ImportError if it's not installed.
    """
    import openpyxl

    if not path or not os.path.isfile(path):
        return {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return {}
    sheets: dict[str, list[list[str]]] = {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            sheets[name] = rows
    finally:
        wb.close()
    return sheets


def _load_csv_pair(left_path: str, right_path: str) -> tuple[list[list[str]], list[list[str]]]:
    return _read_csv_rows(left_path), _read_csv_rows(right_path)


def _load_excel_pair(
    left_path: str, right_path: str
) -> tuple[dict[str, list[list[str]]], dict[str, list[list[str]]]]:
    return _read_excel_sheets(left_path), _read_excel_sheets(right_path)


# ---------------------------------------------------------------------------
# DiffTableModel
# ---------------------------------------------------------------------------


class DiffTableModel(QStandardItemModel):
    """A QStandardItemModel that colours cells based on their diff status.

    Each cell stores its status string in ``Qt.UserRole``.  The background
    colour is derived from :data:`_STATUS_COLORS`.
    """

    def __init__(
        self,
        rows: int = 0,
        columns: int = 0,
        parent: Optional[QWidget] = None,
    ) -> None:
        super().__init__(rows, columns, parent)

    # -- helpers -----------------------------------------------------------

    @staticmethod
    def make_item(text: str, status: str) -> QStandardItem:
        """Create a :class:`QStandardItem` with the appropriate background."""
        item = QStandardItem(text)
        item.setEditable(False)
        item.setData(status, Qt.UserRole)
        bg = _STATUS_COLORS.get(status)
        if bg is not None:
            item.setBackground(bg)
        return item


# ---------------------------------------------------------------------------
# TableView
# ---------------------------------------------------------------------------


def _align_rows_by_key(
    left_rows: list[list[str]],
    right_rows: list[list[str]],
    key_index: int,
) -> tuple[list[Optional[list[str]]], list[Optional[list[str]]]]:
    """Pair rows by the value in *key_index* rather than by position.

    Mirrors ``rcompare_core::CsvDiffEngine::with_key_columns``. Returns two
    equal-length lists in which ``None`` marks "no counterpart on this side",
    so a left-only row is reported as left-only instead of being paired with
    an unrelated right-only row and called "different".

    Left order is preserved; right-only rows follow, in their original order.
    Duplicate keys are matched in occurrence order.
    """
    def key_of(row: list[str]) -> str:
        return row[key_index] if key_index < len(row) else ""

    remaining: dict[str, list[list[str]]] = {}
    for row in right_rows:
        remaining.setdefault(key_of(row), []).append(row)

    aligned_left: list[Optional[list[str]]] = []
    aligned_right: list[Optional[list[str]]] = []
    matched: set[int] = set()

    for row in left_rows:
        candidates = remaining.get(key_of(row))
        aligned_left.append(row)
        if candidates:
            partner = candidates.pop(0)
            matched.add(id(partner))
            aligned_right.append(partner)
        else:
            aligned_right.append(None)

    for row in right_rows:
        if id(row) not in matched:
            aligned_left.append(None)
            aligned_right.append(row)

    return aligned_left, aligned_right


class TableView(QWidget):
    """Side-by-side table comparison widget for CSV / Excel data.

    The widget displays a header bar with left/right file paths, a
    :class:`QTabWidget` for multiple sheets (Excel) or a single table
    (CSV), and a summary bar at the bottom.
    """

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)

        self._pending_table_paths: tuple[str, str] | None = None
        self._table_worker: Optional[FunctionWorker] = None
        # Raw rows kept so changing the header/key controls can re-align
        # without re-reading the files.
        self._left_rows: list[list[str]] = []
        self._right_rows: list[list[str]] = []
        self._suppress_alignment_reload = False

        # Header -----------------------------------------------------------
        self._left_path_label = QLabel("(no file loaded)")
        self._right_path_label = QLabel("(no file loaded)")
        self._left_path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self._right_path_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        header_layout = QHBoxLayout()
        header_layout.addWidget(QLabel("<b>Left:</b>"))
        header_layout.addWidget(self._left_path_label, stretch=1)
        header_layout.addWidget(QLabel("<b>Right:</b>"))
        header_layout.addWidget(self._right_path_label, stretch=1)

        # Alignment controls (WI-5.3) --------------------------------------
        # Positional alignment pairs a left-only row with an unrelated
        # right-only row and calls both "different"; one inserted row cascades
        # through the rest of the file. Choosing a key column fixes that.
        self._header_check = QCheckBox("First row is a header")
        self._header_check.setChecked(True)
        self._header_check.setToolTip(
            "Use the first row for column names instead of Col 1, Col 2, ..."
        )
        self._header_check.toggled.connect(self._on_alignment_changed)

        self._key_combo = QComboBox()
        self._key_combo.setMinimumWidth(160)
        self._key_combo.setToolTip(
            "Match rows by the value of this column instead of by position."
        )
        self._key_combo.setAccessibleName("Row alignment key column")
        self._key_combo.addItem("Align by position", "")
        self._key_combo.currentIndexChanged.connect(self._on_alignment_changed)

        align_layout = QHBoxLayout()
        align_layout.addWidget(self._header_check)
        align_layout.addSpacing(12)
        align_layout.addWidget(QLabel("Key column:"))
        align_layout.addWidget(self._key_combo)
        align_layout.addStretch(1)

        # Tab widget -------------------------------------------------------
        self._tab_widget = QTabWidget()

        # Summary bar ------------------------------------------------------
        self._summary_label = QLabel("")
        self._summary_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        # Main layout ------------------------------------------------------
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(4, 4, 4, 4)
        main_layout.addWidget(QLabel("<b>Data Compare</b>"))
        main_layout.addLayout(header_layout)
        main_layout.addLayout(align_layout)
        main_layout.addWidget(self._tab_widget, stretch=1)
        main_layout.addWidget(self._summary_label)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def compare_csv(self, left_path: str, right_path: str) -> None:
        """Read two CSV files and display a side-by-side diff.

        File reads/parsing run on a background :class:`FunctionWorker`
        thread so large spreadsheets don't block the GUI.
        """
        self._set_path_labels(left_path, right_path)
        self._tab_widget.clear()
        self._summary_label.setText("Loading...")
        self._pending_table_paths = (left_path, right_path)

        worker = FunctionWorker(_load_csv_pair, left_path, right_path, parent=self)
        worker.finished_with_result.connect(
            lambda result, paths=(left_path, right_path): self._on_csv_loaded(paths, result)
        )
        worker.error.connect(
            lambda msg, paths=(left_path, right_path): self._on_table_load_error(paths, msg)
        )
        worker.finished.connect(worker.deleteLater)
        self._table_worker = worker
        worker.start()

    def _on_csv_loaded(
        self, paths: tuple[str, str], result: tuple[list[list[str]], list[list[str]]]
    ) -> None:
        if self._pending_table_paths != paths:
            return  # a newer compare_csv()/compare_excel() call superseded this one
        left_rows, right_rows = result
        self._left_rows, self._right_rows = left_rows, right_rows
        self._refresh_key_choices()
        self._render_csv()

    def _render_csv(self) -> None:
        """(Re)build the CSV tab from the cached rows and current alignment."""
        self._tab_widget.clear()
        left_rows, right_rows = self._left_rows, self._right_rows
        headers = self._column_headers(left_rows, right_rows)

        if self._header_check.isChecked():
            left_rows = left_rows[1:]
            right_rows = right_rows[1:]

        key_index = self._key_column_index()
        if key_index is None:
            left_model, right_model, stats = self._build_models(left_rows, right_rows)
        else:
            aligned_left, aligned_right = _align_rows_by_key(
                left_rows, right_rows, key_index
            )
            left_model, right_model, stats = self._build_models(
                aligned_left, aligned_right, allow_gaps=True
            )

        left_model.setHorizontalHeaderLabels(headers)
        right_model.setHorizontalHeaderLabels(headers)
        splitter = self._create_table_pair(left_model, right_model)
        self._tab_widget.addTab(splitter, "CSV")
        self._update_summary(stats)

    def _column_headers(
        self, left_rows: list[list[str]], right_rows: list[list[str]]
    ) -> list[str]:
        """Return column names: the first row if it is a header, else Col N.

        The hardcoded ``Col 1/2/3`` labels made a keyed comparison impossible
        to configure, because no column had a name to select.
        """
        width = max(
            max((len(r) for r in left_rows), default=0),
            max((len(r) for r in right_rows), default=0),
        )
        if self._header_check.isChecked():
            source = left_rows[0] if left_rows else (right_rows[0] if right_rows else [])
            return [
                source[i] if i < len(source) and source[i] else f"Col {i + 1}"
                for i in range(width)
            ]
        return [f"Col {i + 1}" for i in range(width)]

    def _key_column_index(self) -> Optional[int]:
        """Return the selected key column index, or None for positional."""
        data = self._key_combo.currentData()
        return data if isinstance(data, int) else None

    def _refresh_key_choices(self) -> None:
        """Rebuild the key-column choices from the current header row."""
        previous = self._key_combo.currentText()
        self._suppress_alignment_reload = True
        self._key_combo.clear()
        self._key_combo.addItem("Align by position", "")
        for index, name in enumerate(
            self._column_headers(self._left_rows, self._right_rows)
        ):
            self._key_combo.addItem(name, index)
        restored = self._key_combo.findText(previous)
        self._key_combo.setCurrentIndex(max(0, restored))
        self._suppress_alignment_reload = False

    def _on_alignment_changed(self) -> None:
        """Re-align the loaded rows after a header/key control change."""
        if self._suppress_alignment_reload or not (
            self._left_rows or self._right_rows
        ):
            return
        self._refresh_key_choices()
        self._render_csv()

    @property
    def key_column(self) -> Optional[int]:
        """Selected key column index, or None (used by tests)."""
        return self._key_column_index()

    def compare_excel(self, left_path: str, right_path: str) -> None:
        """Read two Excel files and display a tab-per-sheet diff.

        File reads/parsing run on a background :class:`FunctionWorker`
        thread so large workbooks don't block the GUI.
        """
        self._set_path_labels(left_path, right_path)
        self._tab_widget.clear()

        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self._summary_label.setText(
                "openpyxl is not installed. Install it with: pip install openpyxl"
            )
            return

        self._summary_label.setText("Loading...")
        self._pending_table_paths = (left_path, right_path)

        worker = FunctionWorker(_load_excel_pair, left_path, right_path, parent=self)
        worker.finished_with_result.connect(
            lambda result, paths=(left_path, right_path): self._on_excel_loaded(paths, result)
        )
        worker.error.connect(
            lambda msg, paths=(left_path, right_path): self._on_table_load_error(paths, msg)
        )
        worker.finished.connect(worker.deleteLater)
        self._table_worker = worker
        worker.start()

    def _on_excel_loaded(
        self,
        paths: tuple[str, str],
        result: tuple[dict[str, list[list[str]]], dict[str, list[list[str]]]],
    ) -> None:
        if self._pending_table_paths != paths:
            return  # a newer compare_csv()/compare_excel() call superseded this one
        left_sheets, right_sheets = result

        all_sheet_names: list[str] = []
        for name in left_sheets:
            if name not in all_sheet_names:
                all_sheet_names.append(name)
        for name in right_sheets:
            if name not in all_sheet_names:
                all_sheet_names.append(name)

        total_stats: dict[str, int] = {
            STATUS_IDENTICAL: 0,
            STATUS_DIFFERENT: 0,
            STATUS_LEFT_ONLY: 0,
            STATUS_RIGHT_ONLY: 0,
        }

        for sheet_name in all_sheet_names:
            left_rows = left_sheets.get(sheet_name, [])
            right_rows = right_sheets.get(sheet_name, [])
            left_model, right_model, stats = self._build_models(left_rows, right_rows)
            splitter = self._create_table_pair(left_model, right_model)
            self._tab_widget.addTab(splitter, sheet_name)
            for key in total_stats:
                total_stats[key] += stats.get(key, 0)

        self._update_summary(total_stats)

    def _on_table_load_error(self, paths: tuple[str, str], message: str) -> None:
        if self._pending_table_paths != paths:
            return
        self._summary_label.setText(f"Error loading file: {message}")

    def load_from_cli_data(self, diff_data: dict) -> None:
        """Populate the view from pre-computed CLI bridge diff data.

        Expected format::

            {
                "left_path": str,
                "right_path": str,
                "sheets": [
                    {
                        "name": str,
                        "rows": [
                            {
                                "cells": [
                                    {"left": str, "right": str, "status": str}
                                ]
                            }
                        ]
                    }
                ]
            }

        A flat (non-sheet) format is also accepted::

            {
                "path": str,
                "rows": [
                    {
                        "cells": [
                            {"left": str, "right": str, "status": str}
                        ]
                    }
                ]
            }
        """
        self._tab_widget.clear()

        left_path = diff_data.get("left_path", diff_data.get("path", ""))
        right_path = diff_data.get("right_path", "")
        self._set_path_labels(left_path, right_path)

        total_stats: dict[str, int] = {
            STATUS_IDENTICAL: 0,
            STATUS_DIFFERENT: 0,
            STATUS_LEFT_ONLY: 0,
            STATUS_RIGHT_ONLY: 0,
        }

        sheets = diff_data.get("sheets")
        if sheets is None:
            # Flat format -- single table
            rows = diff_data.get("rows", [])
            left_model, right_model, stats = self._build_models_from_cells(rows)
            splitter = self._create_table_pair(left_model, right_model)
            self._tab_widget.addTab(splitter, "Data")
            for key in total_stats:
                total_stats[key] += stats.get(key, 0)
        else:
            for sheet in sheets:
                name = sheet.get("name", "Sheet")
                rows = sheet.get("rows", [])
                left_model, right_model, stats = self._build_models_from_cells(rows)
                splitter = self._create_table_pair(left_model, right_model)
                self._tab_widget.addTab(splitter, name)
                for key in total_stats:
                    total_stats[key] += stats.get(key, 0)

        self._update_summary(total_stats)

    # ------------------------------------------------------------------
    # Model building
    # ------------------------------------------------------------------

    def _build_models(
        self,
        left_rows: list,
        right_rows: list,
        *,
        allow_gaps: bool = False,
    ) -> tuple[DiffTableModel, DiffTableModel, dict[str, int]]:
        """Build left and right DiffTableModels from raw row data.

        With *allow_gaps*, the two lists are already aligned and may contain
        ``None`` entries marking a missing counterpart (see
        :func:`_align_rows_by_key`).

        Returns (left_model, right_model, stats_dict).
        """
        max_rows = max(len(left_rows), len(right_rows))
        left_cols = max((len(r) for r in left_rows if r is not None), default=0)
        right_cols = max((len(r) for r in right_rows if r is not None), default=0)
        max_cols = max(left_cols, right_cols)

        left_model = DiffTableModel(0, max_cols, self)
        right_model = DiffTableModel(0, max_cols, self)

        # Set column headers (Col 1, Col 2, ...)
        headers = [f"Col {i + 1}" for i in range(max_cols)]
        left_model.setHorizontalHeaderLabels(headers)
        right_model.setHorizontalHeaderLabels(headers)

        stats: dict[str, int] = {
            STATUS_IDENTICAL: 0,
            STATUS_DIFFERENT: 0,
            STATUS_LEFT_ONLY: 0,
            STATUS_RIGHT_ONLY: 0,
        }

        for row_idx in range(max_rows):
            left_row_data = left_rows[row_idx] if row_idx < len(left_rows) else None
            right_row_data = right_rows[row_idx] if row_idx < len(right_rows) else None
            if allow_gaps and left_row_data is None and right_row_data is None:
                continue

            left_items: list[QStandardItem] = []
            right_items: list[QStandardItem] = []

            if left_row_data is None:
                # Row only on right side
                assert right_row_data is not None
                for col_idx in range(max_cols):
                    val = right_row_data[col_idx] if col_idx < len(right_row_data) else ""
                    left_items.append(DiffTableModel.make_item("", STATUS_MISSING))
                    right_items.append(DiffTableModel.make_item(val, STATUS_RIGHT_ONLY))
                stats[STATUS_RIGHT_ONLY] += 1
            elif right_row_data is None:
                # Row only on left side
                for col_idx in range(max_cols):
                    val = left_row_data[col_idx] if col_idx < len(left_row_data) else ""
                    left_items.append(DiffTableModel.make_item(val, STATUS_LEFT_ONLY))
                    right_items.append(DiffTableModel.make_item("", STATUS_MISSING))
                stats[STATUS_LEFT_ONLY] += 1
            else:
                # Both rows exist -- compare cell by cell
                row_identical = True
                for col_idx in range(max_cols):
                    left_val = left_row_data[col_idx] if col_idx < len(left_row_data) else ""
                    right_val = right_row_data[col_idx] if col_idx < len(right_row_data) else ""

                    left_has = col_idx < len(left_row_data)
                    right_has = col_idx < len(right_row_data)

                    if left_has and right_has:
                        if left_val == right_val:
                            status = STATUS_IDENTICAL
                        else:
                            status = STATUS_DIFFERENT
                            row_identical = False
                    elif left_has:
                        status = STATUS_LEFT_ONLY
                        row_identical = False
                    elif right_has:
                        status = STATUS_RIGHT_ONLY
                        row_identical = False
                    else:
                        status = STATUS_MISSING

                    left_items.append(DiffTableModel.make_item(left_val, status))
                    right_items.append(DiffTableModel.make_item(right_val, status))

                if row_identical:
                    stats[STATUS_IDENTICAL] += 1
                else:
                    stats[STATUS_DIFFERENT] += 1

            left_model.appendRow(left_items)
            right_model.appendRow(right_items)

        return left_model, right_model, stats

    def _build_models_from_cells(
        self,
        rows: list[dict],
    ) -> tuple[DiffTableModel, DiffTableModel, dict[str, int]]:
        """Build models from pre-computed cell diffs (CLI bridge format).

        Each row dict has a ``cells`` key containing a list of cell dicts
        with ``left``, ``right``, and ``status`` keys.
        """
        max_cols = max((len(r.get("cells", [])) for r in rows), default=0)

        left_model = DiffTableModel(0, max_cols, self)
        right_model = DiffTableModel(0, max_cols, self)

        headers = [f"Col {i + 1}" for i in range(max_cols)]
        left_model.setHorizontalHeaderLabels(headers)
        right_model.setHorizontalHeaderLabels(headers)

        stats: dict[str, int] = {
            STATUS_IDENTICAL: 0,
            STATUS_DIFFERENT: 0,
            STATUS_LEFT_ONLY: 0,
            STATUS_RIGHT_ONLY: 0,
        }

        for row_dict in rows:
            cells = row_dict.get("cells", [])
            left_items: list[QStandardItem] = []
            right_items: list[QStandardItem] = []
            row_identical = True

            for col_idx in range(max_cols):
                if col_idx < len(cells):
                    cell = cells[col_idx]
                    left_val = str(cell.get("left", ""))
                    right_val = str(cell.get("right", ""))
                    status = cell.get("status", STATUS_IDENTICAL)
                else:
                    left_val = ""
                    right_val = ""
                    status = STATUS_MISSING

                if status != STATUS_IDENTICAL:
                    row_identical = False

                left_items.append(DiffTableModel.make_item(left_val, status))
                right_items.append(DiffTableModel.make_item(right_val, status))

            if row_identical:
                stats[STATUS_IDENTICAL] += 1
            elif all(
                cells[i].get("status") == STATUS_LEFT_ONLY
                for i in range(len(cells))
                if i < len(cells)
            ):
                stats[STATUS_LEFT_ONLY] += 1
            elif all(
                cells[i].get("status") == STATUS_RIGHT_ONLY
                for i in range(len(cells))
                if i < len(cells)
            ):
                stats[STATUS_RIGHT_ONLY] += 1
            else:
                stats[STATUS_DIFFERENT] += 1

            left_model.appendRow(left_items)
            right_model.appendRow(right_items)

        return left_model, right_model, stats

    # ------------------------------------------------------------------
    # Widget helpers
    # ------------------------------------------------------------------

    def _create_table_pair(
        self,
        left_model: DiffTableModel,
        right_model: DiffTableModel,
    ) -> QSplitter:
        """Create a side-by-side QSplitter with two QTableViews."""
        left_table = self._create_table(left_model)
        right_table = self._create_table(right_model)

        # Synchronise vertical scrolling
        left_vbar = left_table.verticalScrollBar()
        right_vbar = right_table.verticalScrollBar()
        syncing = {"flag": False}

        def _on_left_scrolled(value: int) -> None:
            if syncing["flag"]:
                return
            syncing["flag"] = True
            try:
                if right_vbar is not None:
                    right_vbar.setValue(value)
            finally:
                syncing["flag"] = False

        def _on_right_scrolled(value: int) -> None:
            if syncing["flag"]:
                return
            syncing["flag"] = True
            try:
                if left_vbar is not None:
                    left_vbar.setValue(value)
            finally:
                syncing["flag"] = False

        if left_vbar is not None and right_vbar is not None:
            left_vbar.valueChanged.connect(_on_left_scrolled)
            right_vbar.valueChanged.connect(_on_right_scrolled)

        # Synchronise horizontal scrolling
        left_hbar = left_table.horizontalScrollBar()
        right_hbar = right_table.horizontalScrollBar()
        h_syncing = {"flag": False}

        def _on_left_h_scrolled(value: int) -> None:
            if h_syncing["flag"]:
                return
            h_syncing["flag"] = True
            try:
                if right_hbar is not None:
                    right_hbar.setValue(value)
            finally:
                h_syncing["flag"] = False

        def _on_right_h_scrolled(value: int) -> None:
            if h_syncing["flag"]:
                return
            h_syncing["flag"] = True
            try:
                if left_hbar is not None:
                    left_hbar.setValue(value)
            finally:
                h_syncing["flag"] = False

        if left_hbar is not None and right_hbar is not None:
            left_hbar.valueChanged.connect(_on_left_h_scrolled)
            right_hbar.valueChanged.connect(_on_right_h_scrolled)

        # Labels
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.addWidget(QLabel("<b>Left</b>"))
        left_layout.addWidget(left_table)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.addWidget(QLabel("<b>Right</b>"))
        right_layout.addWidget(right_table)

        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)

        return splitter

    @staticmethod
    def _create_table(model: DiffTableModel) -> QTableView:
        """Build and configure a QTableView for data display."""
        table = QTableView()
        table.setModel(model)
        table.setAlternatingRowColors(False)
        table.setSelectionMode(QTableView.NoSelection)
        table.setShowGrid(True)
        table.verticalHeader().setVisible(True)

        header = table.horizontalHeader()
        if header is not None:
            header.setStretchLastSection(False)
            header.setSectionResizeMode(QHeaderView.ResizeToContents)

        return table

    def _set_path_labels(self, left_path: str, right_path: str) -> None:
        """Update the header path labels."""
        left_name = os.path.basename(left_path) if left_path else "(none)"
        right_name = os.path.basename(right_path) if right_path else "(none)"
        self._left_path_label.setText(left_name)
        self._left_path_label.setToolTip(left_path)
        self._right_path_label.setText(right_name)
        self._right_path_label.setToolTip(right_path)

    def _update_summary(self, stats: dict[str, int]) -> None:
        """Update the summary bar with row statistics."""
        identical = stats.get(STATUS_IDENTICAL, 0)
        different = stats.get(STATUS_DIFFERENT, 0)
        left_only = stats.get(STATUS_LEFT_ONLY, 0)
        right_only = stats.get(STATUS_RIGHT_ONLY, 0)
        self._summary_label.setText(
            f"{identical} identical rows, {different} different rows, "
            f"{left_only} left-only, {right_only} right-only"
        )
